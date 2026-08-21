"""
Fill in all the data columns in Episodic Pivots V3.xlsx against a passing-only Benzinga
candidate list (default: Files/EP/Data Pulls/Benzinga EP Candidates - PASSING ONLY.xlsx,
override with --input for a different candidate-list run, e.g. a full-history pull).

Outputs to a standalone file in Data Pulls/ (never writes directly into the V3 file --
same reasoning as everywhere else in this project: keep working files that have pivot
tables/manual structure safe from openpyxl round-trip corruption). Row-aligned to the
V3 sheet's ticker+reaction_date order so it can be pasted in directly.

All forward/outcome metrics (1M/3M/6M High%, Close Performance%) are measured from the
GAP-DAY OPEN, matching the project's MFE definition:
    1M High % = highest high from the gap day through 1 calendar month after, / gap-day open - 1
Bucket/category boundaries are taken directly from the user's master findings doc
(Appendix A/B and Section 6), not guessed.
"""

import argparse
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone

import openpyxl
import pandas as pd
import requests
import zoneinfo
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from requests.adapters import HTTPAdapter
from tqdm import tqdm

sys.path.insert(0, os.path.dirname(__file__))
from build_v2_features import clean_ticker, resolve_historical_ticker, PLAN_CUTOFF

load_dotenv()
POLYGON_API_KEY = os.environ.get("POLYGON_API_KEY")
if not POLYGON_API_KEY:
    sys.exit("POLYGON_API_KEY not found. Check your .env file.")

CANDIDATE_LIST = os.path.join("Files", "EP", "Data Pulls", "Benzinga EP Candidates - PASSING ONLY.xlsx")
OUTPUT_DIR = os.path.join("Files", "EP", "Data Pulls")

BASE_URL = "https://api.polygon.io"
MAX_RETRIES = 5
MAX_WORKERS = 12
ET = zoneinfo.ZoneInfo("America/New_York")
CANDLE_WINDOWS = [1, 5, 10, 15, 30, 60]
TODAY = date.today()

SESSION = requests.Session()
SESSION.mount("https://", HTTPAdapter(pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS))


def api_get(path, params=None):
    params = dict(params or {})
    params["apiKey"] = POLYGON_API_KEY
    url = f"{BASE_URL}{path}"
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = SESSION.get(url, params=params, timeout=30)
        except requests.exceptions.RequestException:
            time.sleep(2 ** attempt)
            continue
        if resp.status_code == 429:
            time.sleep(2 ** attempt)
            continue
        if resp.status_code in (403, 404):
            return None
        resp.raise_for_status()
        return resp.json()
    return None


# ---------------------------------------------------------------------------
# Data pulls
# ---------------------------------------------------------------------------

def get_daily_bars(ticker, from_date, to_date):
    data = api_get(f"/v2/aggs/ticker/{ticker}/range/1/day/{from_date}/{to_date}",
                    {"adjusted": "true", "sort": "asc", "limit": 50000})
    if data is None or not data.get("results"):
        return None
    df = pd.DataFrame(data["results"])
    df["date"] = df["t"].apply(lambda ms: datetime.fromtimestamp(ms / 1000, tz=timezone.utc).date())
    return df[["date", "o", "h", "l", "c", "v", "vw"]].rename(
        columns={"o": "open", "h": "high", "l": "low", "c": "close", "v": "volume", "vw": "vwap"}
    ).sort_values("date").reset_index(drop=True)


def get_minute_bars(ticker, event_date):
    d = event_date.isoformat()
    data = api_get(f"/v2/aggs/ticker/{ticker}/range/1/minute/{d}/{d}", {"adjusted": "true", "sort": "asc", "limit": 1000})
    if data is None or not data.get("results"):
        return None
    df = pd.DataFrame(data["results"])
    df["dt_et"] = df["t"].apply(lambda ms: datetime.fromtimestamp(ms / 1000, tz=timezone.utc).astimezone(ET))
    return df[["dt_et", "c", "v", "vw"]].rename(columns={"c": "close", "v": "volume", "vw": "vwap"})


_ipo_cache, _sic_cache, _shares_cache = {}, {}, {}


def get_ticker_reference(ticker):
    if ticker in _sic_cache:
        return _sic_cache[ticker]
    data = api_get(f"/v3/reference/tickers/{ticker}")
    res = data.get("results") if data else None
    _sic_cache[ticker] = res
    return res


def get_ipo_date(ticker):
    if ticker in _ipo_cache:
        return _ipo_cache[ticker]
    res = get_ticker_reference(ticker)
    list_date = res.get("list_date") if res else None
    if not list_date:
        search = api_get("/v3/reference/tickers", {"ticker": ticker, "active": "false", "limit": 1})
        if search and search.get("results"):
            delisted_utc = search["results"][0].get("delisted_utc")
            if delisted_utc:
                as_of = (datetime.fromisoformat(delisted_utc.replace("Z", "+00:00")).date() - timedelta(days=30)).isoformat()
                data2 = api_get(f"/v3/reference/tickers/{ticker}", {"date": as_of})
                if data2 and data2.get("results"):
                    list_date = data2["results"].get("list_date")
    _ipo_cache[ticker] = list_date
    return list_date


def get_financials(ticker, timeframe, limit, before_date=None):
    params = {"ticker": ticker, "timeframe": timeframe, "limit": limit, "sort": "filing_date", "order": "desc"}
    if before_date is not None:
        params["filing_date.lt"] = before_date.isoformat()
    data = api_get("/vX/reference/financials", params)
    if not data or not data.get("results"):
        return []
    out = []
    for r in data["results"]:
        inc = r.get("financials", {}).get("income_statement", {})
        eps = inc.get("diluted_earnings_per_share", {}).get("value")
        if eps is None:
            eps = inc.get("basic_earnings_per_share", {}).get("value")
        rev = inc.get("revenues", {}).get("value")
        end_date = r.get("end_date")
        if end_date is None:
            continue
        out.append({"end_date": date.fromisoformat(end_date), "eps": eps, "revenue": rev})
    return sorted(out, key=lambda x: x["end_date"], reverse=True)


def get_dividends(ticker, before_date=None):
    params = {"ticker": ticker, "limit": 20, "sort": "ex_dividend_date", "order": "desc"}
    if before_date is not None:
        params["ex_dividend_date.lt"] = before_date.isoformat()
    data = api_get("/v3/reference/dividends", params)
    if not data or not data.get("results"):
        return []
    out = []
    for d in data["results"]:
        exd = d.get("ex_dividend_date")
        if exd:
            out.append({"ex_date": date.fromisoformat(exd), "amount": d.get("cash_amount", 0) or 0})
    return out


def get_benzinga_previous(ticker, release_date):
    data = api_get("/benzinga/v1/earnings", {"ticker": ticker, "date": release_date.isoformat(), "date_status": "confirmed", "limit": 5})
    if not data or not data.get("results"):
        return None, None
    r = data["results"][0]
    return r.get("previous_eps"), r.get("previous_revenue")


# ---------------------------------------------------------------------------
# SIC -> sector mapping (standard SIC division ranges)
# ---------------------------------------------------------------------------

def sic_to_sector(sic_code):
    if not sic_code:
        return "N/A"
    try:
        n = int(sic_code)
    except (TypeError, ValueError):
        return "N/A"
    if 100 <= n <= 999: return "Agriculture, Forestry, Fishing"
    if 1000 <= n <= 1499: return "Mining"
    if 1500 <= n <= 1799: return "Construction"
    if 2000 <= n <= 3999: return "Manufacturing"
    if 4000 <= n <= 4999: return "Transportation & Utilities"
    if 5000 <= n <= 5199: return "Wholesale Trade"
    if 5200 <= n <= 5999: return "Retail Trade"
    if 6000 <= n <= 6799: return "Finance, Insurance, Real Estate"
    if 7000 <= n <= 8999: return "Services"
    if 9100 <= n <= 9999: return "Public Administration"
    return "N/A"


# ---------------------------------------------------------------------------
# Category / bucket classifiers (boundaries taken directly from the master findings doc)
# ---------------------------------------------------------------------------

def na(v):
    return v if v is not None else "N/A"


def bucket_adr(v):
    if v is None: return "N/A"
    if v < 3: return "01 | 2-3%"
    if v < 4: return "02 | 3-4%"
    if v < 5: return "03 | 4-5%"
    if v < 7: return "04 | 5-7%"
    if v < 10: return "05 | 7-10%"
    return "06 | 10%+"


def bucket_mktcap(v):
    if v is None: return "N/A"
    b = v / 1e9
    if b < 1: return "01 | <$1B"
    if b < 2: return "02 | $1-2B"
    if b < 5: return "03 | $2-5B"
    if b < 10: return "04 | $5-10B"
    if b < 25: return "05 | $10-25B"
    if b < 100: return "06 | $25-100B"
    return "07 | $100B+"


def bucket_ath(gap_open, ath_price):
    if ath_price is None or ath_price <= 0 or gap_open is None or gap_open >= ath_price:
        return "01 | At ATH"
    decline = (ath_price - gap_open) / ath_price * 100
    if decline < 10: return "02 | 0-10% below"
    if decline < 25: return "03 | 10-25% below"
    if decline < 50: return "04 | 25-50% below"
    if decline < 70: return "05 | 50-70% below"
    if decline < 85: return "06 | 70-85% below"
    return "07 | 85%+ below"


def bucket_gap(v):
    if v is None: return "N/A"
    if v < 7.5: return "01 | 5-7.5%"
    if v < 10: return "02 | 7.5-10%"
    if v < 15: return "03 | 10-15%"
    if v < 20: return "04 | 15-20%"
    if v < 30: return "05 | 20-30%"
    return "06 | 30%+"


def bucket_revenue_surprise(v):
    if v is None: return "N/A"
    if v < 0: return "01 | Miss (<0%)"
    if v < 0.5: return "02 | 0-0.5%"
    if v < 2.25: return "03 | 0.5-2.25%"
    if v < 4.5: return "04 | 2.25-4.5%"
    if v < 8.5: return "05 | 4.5-8.5%"
    if v < 15: return "06 | 8.5-15%"
    return "07 | 15%+"


def bucket_eps_surprise(v):
    if v is None: return "N/A"
    if v < 0: return "01 | Miss (<0%)"
    if v < 5: return "02 | 0-5%"
    if v < 15: return "03 | 5-15%"
    if v < 30: return "04 | 15-30%"
    if v < 70: return "05 | 30-70%"
    if v < 150: return "06 | 70-150%"
    return "07 | 150%+"


def bucket_revenue_yoy(v):
    if v is None: return "N/A"
    if v < 0: return "01 | Negative"
    if v < 10: return "02 | 0-10%"
    if v < 20: return "03 | 10-20%"
    if v < 35: return "04 | 20-35%"
    if v < 100: return "05 | 35-100%"
    return "06 | 100%+"


def eps_yoy_state(current, prior):
    if current is None or prior is None or prior == 0:
        return "N/A"
    if prior > 0 and current < 0:
        return "01 | Profit -> Loss"
    if prior < 0 and current < 0:
        return "02 | Loss Worsening" if current < prior else "03 | Loss Narrowing"
    if prior < 0 and current >= 0:
        return "04 | Loss -> Profit"
    if prior > 0 and current >= 0:
        if current < prior:
            return "05 | Positive EPS Decline"
        growth = (current - prior) / prior * 100
        if growth < 25: return "06 | EPS Growth 0-25%"
        if growth < 50: return "07 | EPS Growth 25-50%"
        if growth < 100: return "08 | EPS Growth 50-100%"
        return "09 | EPS Growth 100%+"
    return "N/A"


def bucket_ipo_age(years):
    if years is None: return "N/A"
    if years < 1: return "01 | <1 year"
    if years < 3: return "02 | 1-3 years"
    if years < 5: return "03 | 3-5 years"
    if years < 10: return "04 | 5-10 years"
    if years < 20: return "05 | 10-20 years"
    return "06 | 20+ years"


def bucket_whole_day_rvol(x):
    if x is None: return "N/A"
    if x < 2: return "01 | <2x"
    if x < 3: return "02 | 2-3x"
    if x < 5: return "03 | 3-5x"
    if x < 8: return "04 | 5-8x"
    return "05 | 8x+"


def bucket_dollarvol(v):
    if v is None: return "N/A"
    m = v / 1e6
    if m < 25: return "01 | $10-25M"
    if m < 50: return "02 | $25-50M"
    if m < 100: return "03 | $50-100M"
    if m < 250: return "04 | $100-250M"
    if m < 1000: return "05 | $250M-1B"
    return "06 | $1B+"


def bucket_turnover(v):
    """Frozen bins per Swing_Long_EP_Master_Findings_CURRENT.md Section 0."""
    if v is None: return "N/A"
    if v < 0.5: return "01 | <0.5%"
    if v < 1: return "02 | 0.5-1%"
    if v < 2: return "03 | 1-2%"
    if v < 5: return "04 | 2-5%"
    if v < 10: return "05 | 5-10%"
    return "06 | 10%+"


def green_red(close, open_):
    if close is None or open_ is None: return "N/A"
    return "Green" if close >= open_ else "Red"


def same_day(high_date_str, reaction_date):
    if not high_date_str or high_date_str == "N/A": return "N/A"
    return "Y" if date.fromisoformat(high_date_str) == reaction_date else "N"


CANDLE_ADV_BUCKETS = {
    1: [(0.01, "01 | <0.01x ADV"), (0.03, "02 | 0.01-0.03x ADV"), (0.05, "03 | 0.03-0.05x ADV"), (0.10, "04 | 0.05-0.10x ADV"), (None, "05 | 0.10x+ ADV")],
    5: [(0.05, "01 | <0.05x ADV"), (0.10, "02 | 0.05-0.10x ADV"), (0.17, "03 | 0.10-0.17x ADV"), (0.30, "04 | 0.17-0.30x ADV"), (None, "05 | 0.30x+ ADV")],
    10: [(0.10, "01 | <0.10x ADV"), (0.18, "02 | 0.10-0.18x ADV"), (0.30, "03 | 0.18-0.30x ADV"), (0.45, "04 | 0.30-0.45x ADV"), (None, "05 | 0.45x+ ADV")],
    15: [(0.15, "01 | <0.15x ADV"), (0.25, "02 | 0.15-0.25x ADV"), (0.35, "03 | 0.25-0.35x ADV"), (0.60, "04 | 0.35-0.60x ADV"), (None, "05 | 0.60x+ ADV")],
    30: [(0.25, "01 | <0.25x ADV"), (0.40, "02 | 0.25-0.40x ADV"), (0.60, "03 | 0.40-0.60x ADV"), (0.90, "04 | 0.60-0.90x ADV"), (None, "05 | 0.90x+ ADV")],
    60: [(0.40, "01 | <0.40x ADV"), (0.65, "02 | 0.40-0.65x ADV"), (0.90, "03 | 0.65-0.90x ADV"), (1.40, "04 | 0.90-1.40x ADV"), (None, "05 | 1.40x+ ADV")],
}


def bucket_candle_adv(minutes, x):
    if x is None: return "N/A"
    for threshold, label in CANDLE_ADV_BUCKETS[minutes]:
        if threshold is None or x < threshold:
            return label
    return "N/A"


def bucket_outcome_pct(v):
    """7-bin MFE outcome bucket (Section 2.2 of the master doc)."""
    if v is None: return "N/A"
    pct = v * 100 if abs(v) < 5 else v  # tolerate either fraction or already-percent input
    if pct < 10: return "01 | 0-10%"
    if pct < 30: return "02 | 10-30%"
    if pct < 50: return "03 | 30-50%"
    if pct < 70: return "04 | 50-70%"
    if pct < 90: return "05 | 70-90%"
    if pct < 100: return "06 | 90-100%"
    return "07 | 100%+"


def surprise_combo(eps_surprise_pct, rev_surprise_pct):
    if eps_surprise_pct is None or rev_surprise_pct is None:
        return "N/A"
    eps_big, eps_beat = eps_surprise_pct >= 30, eps_surprise_pct >= 0
    rev_big, rev_beat = rev_surprise_pct >= 15, rev_surprise_pct >= 0
    if eps_big and rev_big: return "01 | Both Big Beats"
    if eps_big and rev_beat: return "02 | Big EPS + Revenue Beat"
    if eps_big and not rev_beat: return "03 | Big EPS Beat + Revenue Miss"
    if eps_beat and rev_big: return "04 | EPS Beat + Big Revenue Beat"
    if eps_beat and rev_beat: return "05 | Both Positive, Neither Big"
    if eps_beat and not rev_beat: return "06 | EPS Beat + Revenue Miss"
    if not eps_beat and rev_big: return "07 | EPS Miss + Big Revenue Beat"
    if not eps_beat and rev_beat: return "08 | EPS Miss + Revenue Beat"
    return "09 | Both Miss"


# ---------------------------------------------------------------------------
# SPY Chillax trend color (computed once, reused for every event)
# ---------------------------------------------------------------------------

def build_spy_series():
    spy = get_daily_bars("SPY", "2000-01-01", TODAY.isoformat())
    spy["ma1"] = spy["close"].rolling(10).mean()
    spy["ma2"] = spy["close"].rolling(20).mean()
    spy["ma1_up"] = spy["ma1"] > spy["ma1"].shift(5)
    spy["ma2_up"] = spy["ma2"] > spy["ma2"].shift(5)

    def classify(row):
        if pd.isna(row["ma1"]) or pd.isna(row["ma2"]) or pd.isna(row["ma1_up"]) or pd.isna(row["ma2_up"]):
            return None
        if row["ma1"] > row["ma2"] and row["ma1_up"] and row["ma2_up"]:
            return "01 | Green"
        if row["ma1"] > row["ma2"] and row["ma1_up"] and not row["ma2_up"]:
            return "02 | Light Green"
        if row["ma1"] > row["ma2"] and not row["ma1_up"] and not row["ma2_up"]:
            return "03 | Yellow"
        return "04 | Downtrend"

    spy["color"] = spy.apply(classify, axis=1)
    return spy


def spy_color_for(spy_df, event_date):
    idx = spy_df.index[spy_df["date"] <= event_date]
    if len(idx) == 0:
        return "N/A"
    return na(spy_df.loc[idx[-1], "color"])


# ---------------------------------------------------------------------------
# Per-event feature computation
# ---------------------------------------------------------------------------

def _sum4(periods, field):
    if len(periods) < 4:
        return None
    vals = [p[field] for p in periods[:4]]
    return sum(vals) if all(v is not None for v in vals) else None


def compute_ttm_growth(ticker, event_date):
    # Q4 is often absent from the "quarterly" bucket (folded into the annual filing instead),
    # so annual and quarterly are kept as SEPARATE sequences and never summed together --
    # mixing them would quadruple-count whichever fiscal year's Q4 got folded into an annual total.
    quarterly = get_financials(ticker, "quarterly", 16, before_date=event_date)
    annual = get_financials(ticker, "annual", 6, before_date=event_date)
    quarterly = [p for p in quarterly if p["end_date"] < event_date]
    annual = [p for p in annual if p["end_date"] < event_date]
    if not quarterly and not annual:
        return None, None

    latest_q_end = quarterly[0]["end_date"] if quarterly else date.min
    latest_a_end = annual[0]["end_date"] if annual else date.min

    if latest_a_end > latest_q_end:
        # Most recent completed period is an annual filing -- its own total IS the TTM
        # as of that date, and the prior-year TTM is simply the previous annual filing.
        if len(annual) < 2:
            return None, None
        eps_now, eps_prior = annual[0]["eps"], annual[1]["eps"]
        rev_now, rev_prior = annual[0]["revenue"], annual[1]["revenue"]
    else:
        # Most recent completed period is a quarter -- sum 4 consecutive quarterly-only
        # entries for "now", and the next 4 back for "prior year".
        eps_now, rev_now = _sum4(quarterly, "eps"), _sum4(quarterly, "revenue")
        eps_prior, rev_prior = _sum4(quarterly[4:], "eps"), _sum4(quarterly[4:], "revenue")

    eps_growth = (eps_now - eps_prior) / abs(eps_prior) * 100 if eps_now is not None and eps_prior not in (None, 0) else None
    rev_growth = (rev_now - rev_prior) / abs(rev_prior) * 100 if rev_now is not None and rev_prior not in (None, 0) else None
    return eps_growth, rev_growth


def compute_dividend_yield(ticker, event_date, pre_gap_close):
    divs = get_dividends(ticker, before_date=event_date)
    lo = event_date - timedelta(days=365)
    ttm = sum(d["amount"] for d in divs if lo <= d["ex_date"] < event_date)
    if not pre_gap_close:
        return None
    return (ttm / pre_gap_close) * 100


def compute_candles(minute_df, event_date):
    out = {}
    if minute_df is None or minute_df.empty:
        for m in CANDLE_WINDOWS:
            out[f"close_{m}m"], out[f"vol_{m}m"], out[f"dvol_{m}m"] = None, None, None
        return out
    session_open = datetime(event_date.year, event_date.month, event_date.day, 9, 30, tzinfo=ET)
    regular = minute_df[minute_df["dt_et"] >= session_open].sort_values("dt_et")
    for m in CANDLE_WINDOWS:
        window_end = session_open + timedelta(minutes=m)
        window = regular[regular["dt_et"] < window_end]
        if window.empty:
            out[f"close_{m}m"], out[f"vol_{m}m"], out[f"dvol_{m}m"] = None, None, None
        else:
            out[f"close_{m}m"] = window["close"].iloc[-1]
            out[f"vol_{m}m"] = window["volume"].sum()
            out[f"dvol_{m}m"] = (window["volume"] * window["vwap"]).sum()
    return out


def process_event(e, spy_df):
    ticker = e["ticker"]
    event_date = e["reaction_date"]

    lo = max(event_date - relativedelta(years=2), PLAN_CUTOFF)
    hi = min(TODAY, event_date + relativedelta(months=7))
    bars = get_daily_bars(ticker, lo.isoformat(), hi.isoformat())
    resolved_ticker = ticker
    if bars is None:
        alt = resolve_historical_ticker(ticker, event_date)
        if alt != ticker:
            bars = get_daily_bars(alt, lo.isoformat(), hi.isoformat())
            resolved_ticker = alt
    if bars is None:
        return {"error": True}

    idx = bars.index[bars["date"] == event_date]
    if len(idx) == 0:
        return {"error": True}
    i = idx[0]
    prior = bars.iloc[:i]
    future = bars.iloc[i:]  # includes the gap day itself, per the MFE definition

    gap_open = bars["open"].iloc[i]
    gap_close = bars["close"].iloc[i]
    pre_gap_close = prior["close"].iloc[-1] if not prior.empty else None

    row = {"resolved_ticker": resolved_ticker, "gap_open": gap_open, "gap_close": gap_close}

    # --- Prior ATH ---
    if not prior.empty:
        ath_idx = prior["high"].idxmax()
        ath_price, ath_date = prior.loc[ath_idx, "high"], prior.loc[ath_idx, "date"]
        row["ath_price"] = 0 if gap_open >= ath_price else ath_price
        row["ath_date"] = None if gap_open >= ath_price else ath_date.isoformat()
        row["pct_from_ath"] = 0 if gap_open >= ath_price else (gap_open / ath_price - 1)
        row["_ath_for_bucket"] = ath_price  # raw, for bucketing regardless of the 0-sentinel
    else:
        row["ath_price"], row["ath_date"], row["pct_from_ath"], row["_ath_for_bucket"] = None, None, None, None

    # --- IPO date / age ---
    ipo = get_ipo_date(resolved_ticker)
    row["ipo_date"] = ipo
    if ipo:
        row["ipo_age_years"] = (event_date - date.fromisoformat(ipo)).days / 365.25
    else:
        row["ipo_age_years"] = None

    # --- Pre-gap volume (30D / 100D share + dollar) ---
    for window, label in [(30, "30"), (100, "100")]:
        p = prior.tail(window)
        if len(p) == window and pre_gap_close:
            row[f"svol{label}"] = p["volume"].mean()
            row[f"dvol{label}"] = pre_gap_close * row[f"svol{label}"]
        else:
            row[f"svol{label}"], row[f"dvol{label}"] = None, None

    # --- Gap-day volume / whole-day RVOL ---
    row["gap_vol"] = bars["volume"].iloc[i]
    row["gap_dvol"] = bars["volume"].iloc[i] * bars["vwap"].iloc[i]
    row["whole_day_rvol"] = (row["gap_vol"] / row["svol30"]) if row.get("svol30") else None

    # --- Forward highs/closes, measured from GAP-DAY OPEN ---
    for m in [1, 3, 6]:
        window_end = event_date + relativedelta(months=m)
        complete = window_end <= TODAY
        window_bars = future[future["date"] <= window_end]
        if window_bars.empty:
            row[f"high_{m}m"], row[f"high_{m}m_date"], row[f"high_{m}m_pct"] = None, None, None
        else:
            peak_idx = window_bars["high"].idxmax()
            row[f"high_{m}m"] = window_bars.loc[peak_idx, "high"]
            row[f"high_{m}m_date"] = window_bars.loc[peak_idx, "date"].isoformat()
            row[f"high_{m}m_pct"] = (row[f"high_{m}m"] / gap_open - 1) if gap_open else None
        if not complete:
            row[f"close_{m}m"], row[f"close_{m}m_date"], row[f"close_{m}m_pct"] = None, None, None
        else:
            close_bars = future[future["date"] <= window_end]
            if close_bars.empty:
                row[f"close_{m}m"], row[f"close_{m}m_date"], row[f"close_{m}m_pct"] = None, None, None
            else:
                last = close_bars.iloc[-1]
                row[f"close_{m}m"] = last["close"]
                row[f"close_{m}m_date"] = last["date"].isoformat()
                row[f"close_{m}m_pct"] = (last["close"] / gap_open - 1) if gap_open else None

    # --- TTM growth ---
    row["ttm_eps_growth"], row["ttm_rev_growth"] = compute_ttm_growth(resolved_ticker, event_date)

    # --- Dividend yield ---
    row["div_yield"] = compute_dividend_yield(resolved_ticker, event_date, pre_gap_close)

    # --- 4-week price change (20 trading days) ---
    p20 = prior.tail(20)
    row["price_chg_4w"] = (pre_gap_close / p20["close"].iloc[0] - 1) * 100 if len(p20) == 20 and pre_gap_close else None

    # --- Sector / industry ---
    ref = get_ticker_reference(resolved_ticker)
    row["sic_desc"] = ref.get("sic_description") if ref else None
    row["sector"] = sic_to_sector(ref.get("sic_code")) if ref else None

    # --- Candles ---
    minute_bars = get_minute_bars(resolved_ticker, event_date)
    row.update(compute_candles(minute_bars, event_date))

    # --- Benzinga previous EPS/revenue (for YoY state) ---
    prev_eps, prev_rev = get_benzinga_previous(resolved_ticker, e["release_date"])
    row["prev_eps"], row["prev_rev"] = prev_eps, prev_rev

    row["spy_color"] = spy_color_for(spy_df, event_date)
    return row


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

V3_COLUMNS = [
    "reaction_date", "ticker", "company_name", "SectorCode", "IndCode", "fiscal_period", "fiscal_year",
    "release_time", "release_timing",
    "actual_eps", "estimated_eps", "eps_surprise_percent", "EPS Surprise Category",
    "actual_revenue", "estimated_revenue", "revenue_surprise_percent", "Revenue Surprise % Category",
    "Revenue EPS Surprise Combo", "EPS YoY Category", "Revenue YoY Category", "TTM EPS Growth %", "TTM Revenue Growth %",
    "Dividend Yield %", "4-Week Price Change %",
    "Gap Day Open", "Gap Day Close", "Gap Day Green/Red?", "gap_pct", "Gap % Category", "adr14", "ADR Category",
    "pre_gap_market_cap", "Mkt Cap Category", "avg_share_volume_30d", "dollar_volume_proxy_30d",
    "Pre-Gap 30D Avg Dollar Volume", "Pre-Gap 30D Avg Share Volume", "Pregap 30D avg Dollar Volume Category",
    "Pre-Gap 100D Avg Dollar Volume", "Pre-Gap 100D Avg Share Volume",
    "Prior ATH Price", "Prior ATH Date", "% from ATH", "ATH Category", "IPO Date", "IPO Date Category",
    "SPY Trend Color",
    "Gap Day Total Volume", "Gap Day Total Dollar Volume", "Relative Volume Multiple (Whole Day vs 30D Avg)",
    "Whole Day Relative Volume Category",
    "1M Candle Close", "1M Candle Green/Red?", "1M Candle Volume", "1M Candle Relative Volume 30D", "1M Candle Dollar Volume", "1M Volume Category",
    "5M Candle Close", "5M Candle Green/Red?", "5M Candle Volume", "5M Candle Relative Volume 30D", "5M Candle Dollar Volume", "5M Volume Category",
    "10M Candle Close", "10M Candle Green/Red?", "10M Candle Volume", "10M Candle Relative Volume 30D", "10M Candle Dollar Volume", "10M Volume Category",
    "15M Candle Close", "15M Candle Green/Red?", "15M Candle Volume", "15M Candle Relative Volume 30D", "15M Candle Dollar Volume", "15M Volume Category",
    "30M Candle Close", "30M Candle Green/Red?", "30M Candle Volume", "30M Candle Relative Volume 30D", "30M Candle Dollar Volume", "30M Volume Category",
    "60M Candle Close", "60M Candle Green/Red?", "60M Candle Volume", "60M Candle Relative Volume 30D", "60M Candle Dollar Volume", "60M Volume Category",
    "1M High", "1M High Date", "1M High %", "1M High % Category", "1M Close", "1M Close Date", "1M Close Performance %",
    "3M High", "3M High Date", "3M High %", "3M High % Category", "3M Close", "3M Close Date", "3M Close Performance %",
    "3M Close Performance Category",
    "6M High", "6M High Date", "6M High %", "6M High % Category", "6M Close", "6M Close Date", "6M Close Performance %",
    "Trading Turnover %", "Trading Turnover % Category",
    "1M High Same Day?", "3M High Same Day?", "6M High Same Day?",
]


def assemble_row(e, r):
    if r.get("error"):
        return {c: ("N/A" if c not in ("reaction_date", "ticker") else e.get(c)) for c in V3_COLUMNS}

    eps_yoy = eps_yoy_state(e["actual_eps"], r.get("prev_eps"))
    rev_yoy_pct = ((e["actual_revenue"] - r["prev_rev"]) / abs(r["prev_rev"]) * 100
                   if r.get("prev_rev") not in (None, 0) and e.get("actual_revenue") is not None else None)
    eps_surprise_pct = e["eps_surprise_percent"] * 100 if e.get("eps_surprise_percent") is not None else None
    rev_surprise_pct = e["revenue_surprise_percent"] * 100 if e.get("revenue_surprise_percent") is not None else None

    out = {
        "reaction_date": e["reaction_date"].isoformat(), "ticker": e["ticker"], "company_name": e.get("company_name"),
        "SectorCode": na(r.get("sector")), "IndCode": na(r.get("sic_desc")),
        "fiscal_period": e.get("fiscal_period"), "fiscal_year": e.get("fiscal_year"),
        "release_time": e.get("release_time"), "release_timing": e.get("release_timing"),

        "actual_eps": na(e.get("actual_eps")), "estimated_eps": na(e.get("estimated_eps")),
        "eps_surprise_percent": na(eps_surprise_pct), "EPS Surprise Category": bucket_eps_surprise(eps_surprise_pct),
        "actual_revenue": na(e.get("actual_revenue")), "estimated_revenue": na(e.get("estimated_revenue")),
        "revenue_surprise_percent": na(rev_surprise_pct), "Revenue Surprise % Category": bucket_revenue_surprise(rev_surprise_pct),
        "Revenue EPS Surprise Combo": surprise_combo(eps_surprise_pct, rev_surprise_pct),
        "EPS YoY Category": eps_yoy, "Revenue YoY Category": bucket_revenue_yoy(rev_yoy_pct),
        "TTM EPS Growth %": na(r.get("ttm_eps_growth")), "TTM Revenue Growth %": na(r.get("ttm_rev_growth")),

        "Dividend Yield %": na(r.get("div_yield")), "4-Week Price Change %": na(r.get("price_chg_4w")),

        "Gap Day Open": na(r.get("gap_open")), "Gap Day Close": na(r.get("gap_close")),
        "Gap Day Green/Red?": green_red(r.get("gap_close"), r.get("gap_open")),
        "gap_pct": na(e.get("gap_pct")), "Gap % Category": bucket_gap(e.get("gap_pct")),
        "adr14": na(e.get("adr14")), "ADR Category": bucket_adr(e.get("adr14")),
        "pre_gap_market_cap": na(e.get("pre_gap_market_cap")), "Mkt Cap Category": bucket_mktcap(e.get("pre_gap_market_cap")),
        "avg_share_volume_30d": na(e.get("avg_share_volume_30d")), "dollar_volume_proxy_30d": na(e.get("dollar_volume_proxy_30d")),
        "Pre-Gap 30D Avg Dollar Volume": na(r.get("dvol30")), "Pre-Gap 30D Avg Share Volume": na(r.get("svol30")),
        "Pregap 30D avg Dollar Volume Category": bucket_dollarvol(r.get("dvol30")),
        "Pre-Gap 100D Avg Dollar Volume": na(r.get("dvol100")), "Pre-Gap 100D Avg Share Volume": na(r.get("svol100")),

        "Prior ATH Price": na(r.get("ath_price")), "Prior ATH Date": na(r.get("ath_date")),
        "% from ATH": na(r.get("pct_from_ath")), "ATH Category": bucket_ath(r.get("gap_open"), r.get("_ath_for_bucket")),
        "IPO Date": na(r.get("ipo_date")), "IPO Date Category": bucket_ipo_age(r.get("ipo_age_years")),

        "SPY Trend Color": r.get("spy_color", "N/A"),

        "Gap Day Total Volume": na(r.get("gap_vol")), "Gap Day Total Dollar Volume": na(r.get("gap_dvol")),
        "Relative Volume Multiple (Whole Day vs 30D Avg)": na(r.get("whole_day_rvol")),
        "Whole Day Relative Volume Category": bucket_whole_day_rvol(r.get("whole_day_rvol")),
    }

    for m in CANDLE_WINDOWS:
        rvol = None
        if r.get(f"vol_{m}m") is not None and e.get("avg_share_volume_30d"):
            rvol = r[f"vol_{m}m"] / e["avg_share_volume_30d"]
        out[f"{m}M Candle Close"] = na(r.get(f"close_{m}m"))
        out[f"{m}M Candle Green/Red?"] = green_red(r.get(f"close_{m}m"), r.get("gap_open"))
        out[f"{m}M Candle Volume"] = na(r.get(f"vol_{m}m"))
        out[f"{m}M Candle Relative Volume 30D"] = na(rvol)
        out[f"{m}M Candle Dollar Volume"] = na(r.get(f"dvol_{m}m"))
        out[f"{m}M Volume Category"] = bucket_candle_adv(m, rvol)

    for m, label in [(1, "1M"), (3, "3M"), (6, "6M")]:
        out[f"{label} High"] = na(r.get(f"high_{m}m"))
        out[f"{label} High Date"] = na(r.get(f"high_{m}m_date"))
        out[f"{label} High %"] = na(r.get(f"high_{m}m_pct"))
        out[f"{label} High % Category"] = bucket_outcome_pct(r.get(f"high_{m}m_pct"))
        out[f"{label} Close"] = na(r.get(f"close_{m}m"))
        out[f"{label} Close Date"] = na(r.get(f"close_{m}m_date"))
        out[f"{label} Close Performance %"] = na(r.get(f"close_{m}m_pct"))
        out[f"{label} High Same Day?"] = same_day(r.get(f"high_{m}m_date"), e["reaction_date"])
    out["3M Close Performance Category"] = bucket_outcome_pct(r.get("close_3m_pct"))

    turnover = (r.get("dvol30") / e["pre_gap_market_cap"] * 100
                if r.get("dvol30") is not None and e.get("pre_gap_market_cap") else None)
    out["Trading Turnover %"] = na(turnover)
    out["Trading Turnover % Category"] = bucket_turnover(turnover)

    return out


def load_candidates(input_path, limit=None):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    events = []
    def to_date(v):
        return v.date() if isinstance(v, datetime) else date.fromisoformat(v)

    for row in ws.iter_rows(min_row=2, values_only=True):
        events.append({
            "ticker": clean_ticker(row[idx["ticker"]]), "company_name": row[idx["company_name"]],
            "release_date": to_date(row[idx["release_date"]]), "release_time": row[idx["release_time"]],
            "release_timing": row[idx["release_timing"]],
            "reaction_date": to_date(row[idx["reaction_date"]]),
            "fiscal_period": row[idx["fiscal_period"]], "fiscal_year": row[idx["fiscal_year"]],
            "actual_eps": row[idx["actual_eps"]], "estimated_eps": row[idx["estimated_eps"]],
            "eps_surprise_percent": row[idx["eps_surprise_percent"]] if isinstance(row[idx["eps_surprise_percent"]], (int, float)) else None,
            "actual_revenue": row[idx["actual_revenue"]], "estimated_revenue": row[idx["estimated_revenue"]],
            "revenue_surprise_percent": row[idx["revenue_surprise_percent"]] if isinstance(row[idx["revenue_surprise_percent"]], (int, float)) else None,
            "gap_pct": row[idx["gap_pct"]], "adr14": row[idx["adr14"]],
            "avg_share_volume_30d": row[idx["avg_share_volume_30d"]], "dollar_volume_proxy_30d": row[idx["dollar_volume_proxy_30d"]],
            "pre_gap_market_cap": row[idx["pre_gap_market_cap"]],
        })
    if limit:
        events = events[:limit]
    return events


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=str, default=CANDIDATE_LIST, help="passing-only candidate list to fill")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--label", type=str, default=None)
    args = parser.parse_args()

    events = load_candidates(args.input, limit=args.limit)
    print(f"{len(events)} candidate events loaded from {args.input}")

    print("Building SPY regime series...")
    spy_df = build_spy_series()

    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_event, e, spy_df): e for e in events}
        with tqdm(total=len(futures), desc="Filling V3 features", unit="event") as pbar:
            for future in as_completed(futures):
                e = futures[future]
                r = future.result()
                results.append(assemble_row(e, r))
                pbar.update(1)

    df = pd.DataFrame(results, columns=V3_COLUMNS)
    df["_sort"] = pd.to_datetime(df["reaction_date"])
    df = df.sort_values("_sort").drop(columns="_sort").reset_index(drop=True)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    tag = f" - {args.label}" if args.label else ""
    out_path = os.path.join(OUTPUT_DIR, f"Episodic Pivots V3 Filled{tag}.xlsx")
    try:
        df.to_excel(out_path, index=False)
        print(f"\nWrote {out_path}")
    except PermissionError:
        print(f"\nCould not write {out_path} (likely open in Excel)")


if __name__ == "__main__":
    main()
