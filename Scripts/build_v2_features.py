"""
Build the full feature set on EP Scan V2.xlsx / "OG No Dupes":
  - Gap Day Open / Close
  - Prior ATH price + date (highest daily High strictly before the event date;
    0 if the gap day's open is already at/above that prior ATH)
  - 1M/3M/6M forward High + High Date (rolling max, event date inclusive)
  - 1M/3M/6M forward Close (single-day close on event_date + N months, snapped
    to the nearest trading day on/before that target)

One Polygon daily-bars pull per ticker (PLAN_CUTOFF -> latest needed date), reused
for every event row and every column above. Falls back to a ticker's historical
symbol (via Polygon's ticker-events history) when the current symbol returns
nothing, same fix as the renamed-ticker pass on the original file.
"""

import argparse
import os
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timezone

import openpyxl
import pandas as pd
import requests
from dateutil.relativedelta import relativedelta
from dotenv import load_dotenv
from requests.adapters import HTTPAdapter
from tqdm import tqdm

load_dotenv()

POLYGON_API_KEY = os.environ.get("POLYGON_API_KEY")
if not POLYGON_API_KEY:
    sys.exit("POLYGON_API_KEY not found. Check your .env file.")

INPUT_XLSX = os.path.join("Files", "EP", "EP Scan V2.xlsx")
SHEET_NAME = "OG No Dupes"

BASE_URL = "https://api.polygon.io"
MAX_RETRIES = 5
MAX_WORKERS = 15
WINDOWS_MONTHS = [1, 3, 6]
TODAY = date.today()
# Polygon plan upgraded 2026-08 from 5yr to 20yr bars history -- this now comfortably
# covers Benzinga's earnings feed floor (2011-05-12), so nothing in the EP pipeline
# should get truncated by plan limits anymore.
PLAN_CUTOFF = TODAY - relativedelta(years=20) + relativedelta(days=1)

SESSION = requests.Session()
_adapter = HTTPAdapter(pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS)
SESSION.mount("https://", _adapter)


def api_get(path, params=None):
    params = dict(params or {})
    params["apiKey"] = POLYGON_API_KEY
    url = f"{BASE_URL}{path}"
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = SESSION.get(url, params=params, timeout=30)
        except requests.exceptions.RequestException:
            time.sleep(2 ** attempt)
            continue
        if resp.status_code == 429:
            time.sleep(2 ** attempt)
            continue
        if resp.status_code in (403, 404):
            return None
        resp.raise_for_status()
        return resp.json()
    return None


def clean_ticker(raw_ticker):
    t = raw_ticker.split("^")[0].strip()
    if "." in t:
        base, suffix = t.rsplit(".", 1)
        if suffix.isdigit():
            t = base
    return t


def get_ticker_events(ticker):
    data = api_get(f"/vX/reference/tickers/{ticker}/events")
    if not data or "results" not in data:
        return None
    return data["results"].get("events", [])


def resolve_historical_ticker(ticker, event_date):
    events = get_ticker_events(ticker)
    lookup_ticker = ticker
    if events is None and ticker.endswith("Q") and len(ticker) > 1:
        base = ticker[:-1]
        events = get_ticker_events(base)
        if events is not None:
            lookup_ticker = base
    if not events:
        return lookup_ticker
    changes = sorted(
        (e["date"], e["ticker_change"]["ticker"])
        for e in events if e.get("type") == "ticker_change"
    )
    candidate = changes[0][1]
    for d, sym in changes:
        if date.fromisoformat(d) <= event_date:
            candidate = sym
        else:
            break
    return candidate


def get_daily_bars(ticker, from_date, to_date):
    path = f"/v2/aggs/ticker/{ticker}/range/1/day/{from_date}/{to_date}"
    data = api_get(path, {"adjusted": "true", "sort": "asc", "limit": 50000})
    if data is None or not data.get("results"):
        return None
    df = pd.DataFrame(data["results"])
    df["date"] = df["t"].apply(lambda ms: datetime.fromtimestamp(ms / 1000, tz=timezone.utc).date())
    return df[["date", "o", "h", "c"]].rename(columns={"o": "open", "h": "high", "c": "close"})


def load_events(limit=None):
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header)}
    events = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        ticker_raw = row[idx["Ticker"]]
        fd = row[idx["Full Date"]]
        if not ticker_raw or not fd:
            continue
        event_date = fd.date() if isinstance(fd, datetime) else date.fromisoformat(str(fd))
        events.append({"row": row_num, "ticker_raw": ticker_raw, "ticker": clean_ticker(ticker_raw),
                        "event_date": event_date})
    if limit:
        events = events[:limit]
    return events


def compute_row(bars, event_date):
    out = {}

    gap_row = bars[bars["date"] == event_date]
    gap_open = gap_row["open"].iloc[0] if not gap_row.empty else None
    gap_close = gap_row["close"].iloc[0] if not gap_row.empty else None
    out["gap_open"], out["gap_close"] = gap_open, gap_close

    prior = bars[bars["date"] < event_date]
    if prior.empty:
        out["ath_price"], out["ath_date"], out["ath_truncated"] = None, None, None
    else:
        idx = prior["high"].idxmax()
        ath_high = prior.loc[idx, "high"]
        ath_date = prior.loc[idx, "date"]
        truncated = prior["date"].min() <= PLAN_CUTOFF
        if gap_open is not None and gap_open >= ath_high:
            out["ath_price"], out["ath_date"] = 0, None
        else:
            out["ath_price"], out["ath_date"] = ath_high, ath_date.isoformat()
        out["ath_truncated"] = truncated

    for m in WINDOWS_MONTHS:
        window_end = event_date + relativedelta(months=m)
        complete = window_end <= TODAY
        truncated = event_date < PLAN_CUTOFF
        window_bars = bars[(bars["date"] >= event_date) & (bars["date"] <= window_end)]
        if window_bars.empty:
            out[f"high_{m}m"], out[f"high_{m}m_date"] = None, None
        else:
            peak_idx = window_bars["high"].idxmax()
            out[f"high_{m}m"] = window_bars.loc[peak_idx, "high"]
            out[f"high_{m}m_date"] = window_bars.loc[peak_idx, "date"].isoformat()
        out[f"{m}m_complete"], out[f"{m}m_truncated"] = complete, truncated

        if not complete:
            out[f"close_{m}m"], out[f"close_{m}m_date"] = None, None
        else:
            # must be on/after the event date -- a gap in the ticker's data right after
            # the event (delisting, symbol swap) must not fall back to stale pre-event bars
            close_bars = bars[(bars["date"] >= event_date) & (bars["date"] <= window_end)]
            if close_bars.empty:
                out[f"close_{m}m"], out[f"close_{m}m_date"] = None, None
            else:
                last = close_bars.iloc[-1]
                out[f"close_{m}m"] = last["close"]
                out[f"close_{m}m_date"] = last["date"].isoformat()

    return out


def process_ticker_events(ticker, ticker_events):
    lo = PLAN_CUTOFF
    hi = min(TODAY, max(e["event_date"] + relativedelta(months=max(WINDOWS_MONTHS)) for e in ticker_events))
    if hi < lo:
        hi = lo

    bars = get_daily_bars(ticker, lo.isoformat(), hi.isoformat())
    used_ticker = ticker

    if bars is None:
        resolved = resolve_historical_ticker(ticker, min(e["event_date"] for e in ticker_events))
        if resolved != ticker:
            bars = get_daily_bars(resolved, lo.isoformat(), hi.isoformat())
            used_ticker = resolved

    results = {}
    for e in ticker_events:
        if bars is None:
            results[e["row"]] = {"resolved_ticker": used_ticker}
        else:
            r = compute_row(bars, e["event_date"])
            r["resolved_ticker"] = used_ticker
            results[e["row"]] = r
    return results


def na(v):
    return v if v is not None else "N/A"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    events = load_events(limit=args.limit)
    by_ticker = {}
    for e in events:
        by_ticker.setdefault(e["ticker"], []).append(e)

    print(f"{len(events)} event rows across {len(by_ticker)} unique tickers")

    all_results = {}
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(process_ticker_events, t, evs): t for t, evs in by_ticker.items()}
        with tqdm(total=len(futures), desc="Building V2 features", unit="ticker") as pbar:
            for future in as_completed(futures):
                all_results.update(future.result())
                pbar.update(1)

    if args.dry_run:
        rows = []
        for e in events:
            r = all_results.get(e["row"], {})
            rows.append({"ticker": e["ticker"], "resolved": r.get("resolved_ticker"),
                         "event_date": e["event_date"].isoformat(),
                         "gap_open": r.get("gap_open"), "gap_close": r.get("gap_close"),
                         "ath_price": r.get("ath_price"), "ath_date": r.get("ath_date"),
                         "high_1m": r.get("high_1m"), "close_1m": r.get("close_1m"),
                         "high_6m": r.get("high_6m"), "close_6m": r.get("close_6m")})
        print(pd.DataFrame(rows).to_string(index=False))
        return

    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=False)
    ws = wb[SHEET_NAME]
    header = [c.value for c in ws[1]]
    next_col = len(header) + 1

    columns = ["Gap Day Open", "Gap Day Close", "Prior ATH Price", "Prior ATH Date", "Prior ATH Data Truncated"]
    for m in WINDOWS_MONTHS:
        columns += [f"{m}M High", f"{m}M High Date", f"{m}M Window Complete", f"{m}M Plan Data Truncated",
                    f"{m}M Close", f"{m}M Close Date"]

    col_map = {}
    for name in columns:
        ws.cell(row=1, column=next_col, value=name)
        col_map[name] = next_col
        next_col += 1

    for e in events:
        r = all_results.get(e["row"], {})
        row_num = e["row"]
        ws.cell(row=row_num, column=col_map["Gap Day Open"], value=na(r.get("gap_open")))
        ws.cell(row=row_num, column=col_map["Gap Day Close"], value=na(r.get("gap_close")))
        ws.cell(row=row_num, column=col_map["Prior ATH Price"], value=na(r.get("ath_price")))
        ws.cell(row=row_num, column=col_map["Prior ATH Date"], value=na(r.get("ath_date")))
        ws.cell(row=row_num, column=col_map["Prior ATH Data Truncated"], value=r.get("ath_truncated"))
        for m in WINDOWS_MONTHS:
            ws.cell(row=row_num, column=col_map[f"{m}M High"], value=na(r.get(f"high_{m}m")))
            ws.cell(row=row_num, column=col_map[f"{m}M High Date"], value=na(r.get(f"high_{m}m_date")))
            ws.cell(row=row_num, column=col_map[f"{m}M Window Complete"], value=r.get(f"{m}m_complete"))
            ws.cell(row=row_num, column=col_map[f"{m}M Plan Data Truncated"], value=r.get(f"{m}m_truncated"))
            ws.cell(row=row_num, column=col_map[f"{m}M Close"], value=na(r.get(f"close_{m}m")))
            ws.cell(row=row_num, column=col_map[f"{m}M Close Date"], value=na(r.get(f"close_{m}m_date")))

    wb.save(INPUT_XLSX)
    print(f"\nWrote {len(events)} rows / {len(columns)} new columns to {INPUT_XLSX}")


if __name__ == "__main__":
    main()
